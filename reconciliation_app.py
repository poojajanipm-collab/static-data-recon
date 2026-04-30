from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill

APP_DIR = Path(__file__).resolve().parent
DEFAULT_PAYROLL_FILENAME = "Payroll_System_Export.csv"
DEFAULT_HR_FILENAME = "HR_Master_System_Export.csv"
STATIC_PAYROLL_PATH = APP_DIR / DEFAULT_PAYROLL_FILENAME
STATIC_HR_PATH = APP_DIR / DEFAULT_HR_FILENAME
LOCAL_PAYROLL_PATH = Path(r"C:\Users\Pooja\Documents\Codex\2026-04-20-files-mentioned-by-the-user-payroll\Static\Payroll_System_Export.csv")
LOCAL_HR_PATH = Path(r"C:\Users\Pooja\Documents\Codex\2026-04-20-files-mentioned-by-the-user-payroll\Static\HR_Master_System_Export.csv")

KEY_COLUMN = "Employee_ID"
COMPARE_FIELDS = [
    "First_Name",
    "Last_Name",
    "Gender",
    "DOB",
    "Department",
    "Salary",
    "Joining_Date",
    "Termination_Date",
    "Status",
    "Phone",
    "Address",
]
DATE_FIELDS = {"DOB", "Joining_Date", "Termination_Date"}
NUMERIC_FIELDS = {"Salary"}
PHONE_FIELDS = {"Phone"}

RED_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_phone(value: object) -> str:
    return re.sub(r"\D", "", normalize_text(value))


def normalize_date(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""

    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return text
    return parsed.strftime("%Y-%m-%d")


def normalize_number(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""

    number = pd.to_numeric(text, errors="coerce")
    if pd.isna(number):
        return text
    return f"{float(number):.2f}"


def normalize_value(field: str, value: object) -> str:
    if field in DATE_FIELDS:
        return normalize_date(value)
    if field in NUMERIC_FIELDS:
        return normalize_number(value)
    if field in PHONE_FIELDS:
        return normalize_phone(value)
    return normalize_text(value)


def default_candidates(file_name: str, local_path: Path) -> list[Path]:
    return [
        APP_DIR / file_name,
        local_path,
    ]


def find_existing_default_path(file_name: str, local_path: Path) -> Path | None:
    for path in default_candidates(file_name, local_path):
        if path.exists():
            return path
    return None


def load_csv(uploaded_file, file_name: str, local_path: Path, label: str) -> tuple[pd.DataFrame, str]:
    if uploaded_file is not None:
        return pd.read_csv(uploaded_file, dtype=str).fillna(""), f"uploaded {label} file"

    default_path = find_existing_default_path(file_name, local_path)
    if default_path is not None:
        return pd.read_csv(default_path, dtype=str).fillna(""), f"default file at {default_path}"

    raise FileNotFoundError(
        f"{label} file was not uploaded and no bundled default file was found. "
        f"Expected one of: {', '.join(str(path) for path in default_candidates(file_name, local_path))}"
    )


def validate_columns(df: pd.DataFrame, label: str) -> None:
    required_columns = [KEY_COLUMN, *COMPARE_FIELDS]
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise ValueError(f"{label} is missing required columns: {', '.join(missing)}")


def prepare_dataframe(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = [column.strip() for column in df.columns]
    validate_columns(df, source_name)
    df[KEY_COLUMN] = df[KEY_COLUMN].astype(str).str.strip()
    df = df.drop_duplicates(subset=[KEY_COLUMN], keep="first")
    return df[[KEY_COLUMN, *COMPARE_FIELDS]]


def build_issue_details(
    status: str,
    mismatched_fields: list[str],
    hr_row: dict[str, object] | None,
    payroll_row: dict[str, object] | None,
) -> str:
    if status == "Missing in HR":
        return "Employee record exists in Payroll file but is missing in HR file."
    if status == "Missing in Payroll":
        return "Employee record exists in HR file but is missing in Payroll file."

    issues: list[str] = []
    for field in mismatched_fields:
        hr_value = "" if hr_row is None else normalize_text(hr_row.get(field, ""))
        payroll_value = "" if payroll_row is None else normalize_text(payroll_row.get(field, ""))
        issues.append(f"{field}: HR='{hr_value}' | Payroll='{payroll_value}'")

    return " ; ".join(issues)


def build_reconciliation(
    hr_df: pd.DataFrame, payroll_df: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    all_employee_ids = sorted(set(hr_df[KEY_COLUMN]).union(set(payroll_df[KEY_COLUMN])))
    hr_lookup = hr_df.set_index(KEY_COLUMN).to_dict("index")
    payroll_lookup = payroll_df.set_index(KEY_COLUMN).to_dict("index")

    summary_rows = []
    detail_rows = []

    for employee_id in all_employee_ids:
        hr_row = hr_lookup.get(employee_id)
        payroll_row = payroll_lookup.get(employee_id)

        mismatched_fields: list[str] = []
        status = "Match"

        if hr_row is None:
            status = "Missing in HR"
        elif payroll_row is None:
            status = "Missing in Payroll"
        else:
            for field in COMPARE_FIELDS:
                hr_value = normalize_value(field, hr_row.get(field, ""))
                payroll_value = normalize_value(field, payroll_row.get(field, ""))
                if hr_value != payroll_value:
                    mismatched_fields.append(field)

            if mismatched_fields:
                status = "Mismatch"

        summary_rows.append(
            {
                KEY_COLUMN: employee_id,
                "Result": status,
                "Mismatch_Count": len(mismatched_fields),
                "Mismatched_Fields": ", ".join(mismatched_fields),
                "Issue_Details": build_issue_details(status, mismatched_fields, hr_row, payroll_row),
            }
        )

        detail_row = {
            KEY_COLUMN: employee_id,
            "Result": status,
            "Mismatched_Fields": ", ".join(mismatched_fields),
            "Issue_Details": build_issue_details(status, mismatched_fields, hr_row, payroll_row),
        }

        for field in COMPARE_FIELDS:
            hr_raw = "" if hr_row is None else normalize_text(hr_row.get(field, ""))
            payroll_raw = "" if payroll_row is None else normalize_text(payroll_row.get(field, ""))

            detail_row[f"HR_{field}"] = hr_raw
            detail_row[f"Payroll_{field}"] = payroll_raw
            detail_row[f"{field}_Match"] = (
                "Missing Record"
                if hr_row is None or payroll_row is None
                else "Yes"
                if normalize_value(field, hr_raw) == normalize_value(field, payroll_raw)
                else "No"
            )

        detail_rows.append(detail_row)

    return pd.DataFrame(summary_rows), pd.DataFrame(detail_rows)


def style_detail_dataframe(detail_df: pd.DataFrame) -> pd.io.formats.style.Styler:
    def highlight_row(row: pd.Series) -> list[str]:
        styles = [""] * len(row)
        row_map = {column: index for index, column in enumerate(detail_df.columns)}
        mismatched_fields = {
            field.strip()
            for field in normalize_text(row["Mismatched_Fields"]).split(",")
            if field.strip()
        }

        if row["Result"] in {"Missing in HR", "Missing in Payroll"}:
            for column, index in row_map.items():
                if column.startswith("HR_") or column.startswith("Payroll_"):
                    styles[index] = "background-color: #f8d7da; color: #c62828; font-weight: 600;"
            for column in ["Result", "Mismatched_Fields", "Issue_Details"]:
                if column in row_map:
                    styles[row_map[column]] = "color: #c62828; font-weight: 700;"
            return styles

        for field in mismatched_fields:
            for column in [f"HR_{field}", f"Payroll_{field}", f"{field}_Match"]:
                if column in row_map:
                    styles[row_map[column]] = "background-color: #f8d7da; color: #c62828; font-weight: 600;"

        for column in ["Result", "Mismatched_Fields", "Issue_Details"]:
            if column in row_map and row["Result"] != "Match":
                styles[row_map[column]] = "color: #c62828; font-weight: 700;"

        return styles

    return detail_df.style.apply(highlight_row, axis=1)


def style_summary_dataframe(summary_df: pd.DataFrame) -> pd.io.formats.style.Styler:
    def highlight_summary_row(row: pd.Series) -> list[str]:
        styles = [""] * len(row)
        row_map = {column: index for index, column in enumerate(summary_df.columns)}

        if row["Result"] != "Match":
            for column in ["Result", "Mismatch_Count", "Mismatched_Fields", "Issue_Details"]:
                if column in row_map:
                    styles[row_map[column]] = "background-color: #f8d7da; color: #c62828; font-weight: 700;"

        return styles

    return summary_df.style.apply(highlight_summary_row, axis=1)


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 28)


def write_dataframe_to_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append(list(row))
    autosize_worksheet(worksheet)


def mark_discrepancies_in_excel(worksheet, detail_df: pd.DataFrame) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}

    for row_idx, row in enumerate(detail_df.itertuples(index=False), start=2):
        result = getattr(row, "Result")
        mismatched = {
            field.strip()
            for field in normalize_text(getattr(row, "Mismatched_Fields")).split(",")
            if field.strip()
        }

        if result in {"Missing in HR", "Missing in Payroll"}:
            for header, column_index in headers.items():
                if str(header).startswith("HR_") or str(header).startswith("Payroll_"):
                    worksheet.cell(row=row_idx, column=column_index).fill = RED_FILL
            continue

        for field in mismatched:
            for header in [f"HR_{field}", f"Payroll_{field}", f"{field}_Match"]:
                column_index = headers.get(header)
                if column_index:
                    worksheet.cell(row=row_idx, column=column_index).fill = RED_FILL


def create_excel_report(summary_df: pd.DataFrame, detail_df: pd.DataFrame) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    write_dataframe_to_sheet(summary_sheet, summary_df)

    detail_sheet = workbook.create_sheet("Detailed_Reconciliation")
    write_dataframe_to_sheet(detail_sheet, detail_df)
    mark_discrepancies_in_excel(detail_sheet, detail_df)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def result_counts(summary_df: pd.DataFrame) -> dict[str, int]:
    counts = summary_df["Result"].value_counts().to_dict()
    return {
        "Match": counts.get("Match", 0),
        "Mismatch": counts.get("Mismatch", 0),
        "Missing in HR": counts.get("Missing in HR", 0),
        "Missing in Payroll": counts.get("Missing in Payroll", 0),
    }


def display_metrics(summary_df: pd.DataFrame) -> None:
    counts = result_counts(summary_df)
    columns = st.columns(4)
    metrics = [
        ("Matches", counts["Match"]),
        ("Mismatches", counts["Mismatch"]),
        ("Missing in HR", counts["Missing in HR"]),
        ("Missing in Payroll", counts["Missing in Payroll"]),
    ]

    for column, (label, value) in zip(columns, metrics):
        column.metric(label, value)


def default_file_message() -> str:
    messages: list[str] = []
    if STATIC_PAYROLL_PATH.exists():
        messages.append(f"Payroll default: `{STATIC_PAYROLL_PATH}`")
    elif LOCAL_PAYROLL_PATH.exists():
        messages.append(f"Payroll local fallback: `{LOCAL_PAYROLL_PATH}`")

    if STATIC_HR_PATH.exists():
        messages.append(f"HR default: `{STATIC_HR_PATH}`")
    elif LOCAL_HR_PATH.exists():
        messages.append(f"HR local fallback: `{LOCAL_HR_PATH}`")

    if messages:
        return " | ".join(messages)
    return (
        "Upload both files, or place "
        f"`{DEFAULT_PAYROLL_FILENAME}` and `{DEFAULT_HR_FILENAME}` in `{APP_DIR}` "
        "to let users run reconciliation without uploading."
    )


def main() -> None:
    st.set_page_config(page_title="Employee Data Reconciliation", layout="wide")
    st.title("Employee Personal Data Reconciliation")
    st.write(
        "Compare HR and Payroll employee records, highlight discrepancies in red, "
        "and download a reconciliation report in Excel."
    )
    st.caption(default_file_message())

    with st.sidebar:
        st.header("Input Files")
        payroll_upload = st.file_uploader("Upload Payroll CSV", type=["csv"], key="payroll")
        hr_upload = st.file_uploader("Upload HR CSV", type=["csv"], key="hr")
        run_button = st.button("Run Reconciliation", type="primary", use_container_width=True)

    if not run_button:
        st.info("Upload files if needed, then click 'Run Reconciliation'.")
        return

    try:
        payroll_raw_df, payroll_source = load_csv(
            payroll_upload,
            DEFAULT_PAYROLL_FILENAME,
            LOCAL_PAYROLL_PATH,
            "Payroll",
        )
        hr_raw_df, hr_source = load_csv(
            hr_upload,
            DEFAULT_HR_FILENAME,
            LOCAL_HR_PATH,
            "HR",
        )
        payroll_df = prepare_dataframe(payroll_raw_df, "Payroll")
        hr_df = prepare_dataframe(hr_raw_df, "HR")
        summary_df, detail_df = build_reconciliation(hr_df, payroll_df)
        excel_report = create_excel_report(summary_df, detail_df)
    except Exception as exc:
        st.error(f"Unable to complete reconciliation: {exc}")
        return

    st.success(f"Using {payroll_source} and {hr_source}.")

    display_metrics(summary_df)

    st.subheader("Summary")
    st.dataframe(style_summary_dataframe(summary_df), use_container_width=True, hide_index=True)

    mismatch_summary = summary_df[summary_df["Result"] != "Match"][
        [KEY_COLUMN, "Result", "Issue_Details"]
    ].copy()
    st.subheader("Mismatch Summary")
    if mismatch_summary.empty:
        st.success("No discrepancies found.")
    else:
        st.dataframe(style_summary_dataframe(mismatch_summary), use_container_width=True, hide_index=True)

    st.subheader("Detailed Comparison")
    st.dataframe(style_detail_dataframe(detail_df), use_container_width=True, hide_index=True)

    mismatches_only = detail_df[detail_df["Result"] != "Match"].copy()
    st.subheader("Discrepancies Only")
    if mismatches_only.empty:
        st.success("No discrepancies found.")
    else:
        st.dataframe(style_detail_dataframe(mismatches_only), use_container_width=True, hide_index=True)

    st.download_button(
        label="Download Excel Report",
        data=excel_report,
        file_name="employee_reconciliation_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
